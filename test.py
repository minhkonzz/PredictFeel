from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import numpy as np
import pandas as pd
import statistics
from sklearn.model_selection import train_test_split

class Ui_App(object):

  def __init__(self):
    self.output = ""

  def setupUi(self, App): 
    App.setObjectName("App")
    App.resize(908, 727)
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(11)
    font.setBold(True)
    font.setWeight(75)
    App.setFont(font)
    App.setWindowOpacity(1.0)
    App.setStyleSheet("border-radius: 10px;")
    self.centralwidget = QtWidgets.QWidget(App)
    self.centralwidget.setObjectName("centralwidget")
    self.predict_button = QtWidgets.QPushButton(self.centralwidget)
    self.predict_button.setGeometry(QtCore.QRect(60, 630, 231, 61))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.predict_button.setFont(font)
    self.predict_button.setStyleSheet(
      "border-radius: 14px;\n"
      "background-color: black;\n"
      "color: white;"
    )
    self.predict_button.setObjectName("predict_button")
    self.export_excel_button = QtWidgets.QPushButton(self.centralwidget)
    self.export_excel_button.setGeometry(QtCore.QRect(340, 630, 231, 61))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.export_excel_button.setFont(font)
    self.export_excel_button.setStyleSheet(
      "border-radius: 12px;\n"
      "background-color: black;\n"
      "color: white;"
    )
    self.export_excel_button.setObjectName("export_excel_button")
    self.retype_button = QtWidgets.QPushButton(self.centralwidget)
    self.retype_button.setGeometry(QtCore.QRect(610, 630, 231, 61))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.retype_button.setFont(font)
    self.retype_button.setStyleSheet(
      "border-radius: 12px;\n"
      "background-color: black;\n"
      "color: white;"
    )
    self.retype_button.setObjectName("retype_button")
    self.name_label = QtWidgets.QLabel(self.centralwidget)
    self.name_label.setGeometry(QtCore.QRect(60, 120, 131, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.name_label.setFont(font)
    self.name_label.setObjectName("name_label")
    self.gender_label = QtWidgets.QLabel(self.centralwidget)
    self.gender_label.setGeometry(QtCore.QRect(60, 220, 131, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)  
    self.gender_label.setFont(font)
    self.gender_label.setObjectName("gender_label")
    self.age_label = QtWidgets.QLabel(self.centralwidget)
    self.age_label.setGeometry(QtCore.QRect(60, 300, 131, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.age_label.setFont(font)
    self.age_label.setObjectName("age_label")
    self.job_label = QtWidgets.QLabel(self.centralwidget)
    self.job_label.setGeometry(QtCore.QRect(60, 400, 131, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.job_label.setFont(font)
    self.job_label.setObjectName("job_label")
    self.salary_label = QtWidgets.QLabel(self.centralwidget)
    self.salary_label.setGeometry(QtCore.QRect(60, 500, 131, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(10)
    font.setBold(True)
    font.setWeight(75)
    self.salary_label.setFont(font)
    self.salary_label.setObjectName("salary_label")
    self.gender_cb = QtWidgets.QComboBox(self.centralwidget)
    self.gender_cb.setGeometry(QtCore.QRect(210, 220, 321, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(9)
    font.setBold(True)
    font.setWeight(75)
    self.gender_cb.setFont(font)
    self.gender_cb.setObjectName("gender_cb")
    self.gender_cb.addItem("")
    self.gender_cb.addItem("")
    self.name_textedit = QtWidgets.QTextEdit(self.centralwidget)
    self.name_textedit.setGeometry(QtCore.QRect(210, 120, 631, 51))
    font = QtGui.QFont()
    font.setPointSize(-1)
    self.name_textedit.setFont(font)
    self.name_textedit.setStyleSheet(
      "border-radius: 14px;\n"
      "font-size: 20px;"
    )
    self.name_textedit.setObjectName("name_textedit")
    self.age_textedit = QtWidgets.QTextEdit(self.centralwidget)
    self.age_textedit.setGeometry(QtCore.QRect(210, 300, 631, 51))
    font = QtGui.QFont()
    font.setPointSize(-1)
    self.age_textedit.setFont(font)
    self.age_textedit.setStyleSheet(
      "border-radius: 14px;\n"
      "font-size: 20px;"
    )
    self.age_textedit.setObjectName("age_textedit")
    self.job_textedit = QtWidgets.QTextEdit(self.centralwidget)
    self.job_textedit.setGeometry(QtCore.QRect(210, 400, 631, 51))
    font = QtGui.QFont()
    font.setPointSize(-1)
    self.job_textedit.setFont(font)
    self.job_textedit.setStyleSheet(
      "border-radius: 14px;\n"
      "font-size: 20px;"
    )
    self.job_textedit.setObjectName("job_textedit")
    self.salary_textedit = QtWidgets.QTextEdit(self.centralwidget)
    self.salary_textedit.setGeometry(QtCore.QRect(210, 500, 631, 51))
    font = QtGui.QFont()
    font.setPointSize(-1)
    self.salary_textedit.setFont(font)
    self.salary_textedit.setStyleSheet(
      "border-radius: 14px;\n"
      "font-size: 20px;"
    )
    self.salary_textedit.setObjectName("salary_textedit")
    self.main_label = QtWidgets.QLabel(self.centralwidget)
    self.main_label.setGeometry(QtCore.QRect(160, 30, 631, 51))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(14)
    font.setBold(True)
    font.setWeight(75)
    self.main_label.setFont(font)
    self.main_label.setLayoutDirection(QtCore.Qt.LeftToRight)
    self.main_label.setStyleSheet("")
    self.main_label.setAlignment(QtCore.Qt.AlignCenter)
    self.main_label.setObjectName("main_label")
    self.name_error_label = QtWidgets.QLabel(self.centralwidget)
    self.name_error_label.setGeometry(QtCore.QRect(210, 170, 700, 41))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(9)
    font.setBold(False)
    font.setWeight(50)
    self.name_error_label.setFont(font)
    self.name_error_label.setStyleSheet("color: red;")
    self.name_error_label.setObjectName("name_error_label")
    self.age_error_label = QtWidgets.QLabel(self.centralwidget)
    self.age_error_label.setGeometry(QtCore.QRect(210, 350, 700, 41))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(9)
    font.setBold(False)
    font.setWeight(50)
    self.age_error_label.setFont(font)
    self.age_error_label.setStyleSheet("color: red;")
    self.age_error_label.setObjectName("age_error_label")
    self.job_error_label = QtWidgets.QLabel(self.centralwidget)
    self.job_error_label.setGeometry(QtCore.QRect(210, 450, 700, 41))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(9)
    font.setBold(False)
    font.setWeight(50)
    self.job_error_label.setFont(font)
    self.job_error_label.setStyleSheet("color: red;")
    self.job_error_label.setObjectName("job_error_label")
    self.salary_error_label = QtWidgets.QLabel(self.centralwidget)
    self.salary_error_label.setGeometry(QtCore.QRect(210, 550, 700, 41))
    font = QtGui.QFont()
    font.setFamily("Courier New")
    font.setPointSize(9)
    font.setBold(False)
    font.setWeight(50)
    self.salary_error_label.setFont(font)
    self.salary_error_label.setStyleSheet("color: red;")
    self.salary_error_label.setObjectName("salary_error_label")
    App.setCentralWidget(self.centralwidget)

    self.retranslateUi(App)
    QtCore.QMetaObject.connectSlotsByName(App)

    self.predict_button.clicked.connect(self.show_predict_result)
    self.export_excel_button.clicked.connect(self.export_excel)
    self.retype_button.clicked.connect(self.reset_user_input)

    self.name_textedit.textChanged.connect(self.clear_name_error)
    self.age_textedit.textChanged.connect(self.clear_age_error)
    self.job_textedit.textChanged.connect(self.clear_job_error)
    self.salary_textedit.textChanged.connect(self.clear_salary_error)

  def retranslateUi(self, App):
    _translate = QtCore.QCoreApplication.translate
    App.setWindowTitle(_translate("App", "Main Program"))
    self.predict_button.setText(_translate("App", "D??? ??o??n"))
    self.export_excel_button.setText(_translate("App", "L??u Excel"))
    self.retype_button.setText(_translate("App", "Nh???p l???i"))
    self.name_label.setText(_translate("App", "T??n c???a b???n"))
    self.gender_label.setText(_translate("App", "Gi???i t??nh"))
    self.age_label.setText(_translate("App", "Tu???i"))
    self.job_label.setText(_translate("App", "Ngh??? nghi???p"))
    self.salary_label.setText(_translate("App", "M???c L????ng"))
    self.gender_cb.setItemText(0, _translate("App", "Nam"))
    self.gender_cb.setItemText(1, _translate("App", "N???"))
    self.main_label.setText(_translate("App", "D??? ??o??n ch??? s??? h???nh ph??c d???a v??o m???c l????ng"))
    self.name_error_label.setText(_translate("App", ""))
    self.age_error_label.setText(_translate("App", ""))
    self.job_error_label.setText(_translate("App", ""))
    self.salary_error_label.setText(_translate("App", ""))

  def is_has_error_input(self):
    import re
    is_name_error = is_age_error = is_job_error = is_salary_error = False
    name_pattern = "\s|[a-zA-Z]"
    age_pattern = "\s|[0-9]"
    job_pattern = "\s|\w"
    salary_pattern = "\s|[^A-Za-z]"
    input_name = self.name_textedit.toPlainText()
    input_age = self.age_textedit.toPlainText()
    input_job = self.job_textedit.toPlainText()
    input_salary = self.salary_textedit.toPlainText()
    if len(re.findall(name_pattern, input_name)) < len(input_name):
      self.name_error_label.setText("Ch??? ???????c ch???a k?? t??? ch??? ho???c kho???ng tr???ng")
      is_name_error = True
    if len(re.findall(age_pattern, input_age)) < len(input_age):
      self.age_error_label.setText("Ch??? ???????c ph??p s???")
      is_age_error = True
    if len(re.findall(job_pattern, input_job)) < len(input_job):
      self.job_error_label.setText("Ch??? ???????c ph??p ch???a k?? t??? ch???, s??? ho???c kho???ng tr???ng")
      is_job_error = True
    if len(re.findall(salary_pattern, input_salary)) < len(input_salary):
      self.salary_error_label.setText("Kh??ng cho ph??p k?? t??? ch???")
      is_salary_error = True
    return is_name_error or is_age_error or is_job_error or is_salary_error

  def clear_error(self, error_qlabel):
    if error_qlabel.text():
      error_qlabel.setText("")

  def clear_name_error(self):
    self.clear_error(self.name_error_label)

  def clear_age_error(self):
    self.clear_error(self.age_error_label)

  def clear_job_error(self):
    self.clear_error(self.job_error_label)

  def clear_salary_error(self):
    self.clear_error(self.salary_error_label)

  def coreProcess(self):
    from sklearn.linear_model import LinearRegression
    dataframe = pd.read_csv("income.data.csv")
    X = "income"
    Y = "happiness"
    x_income = dataframe[X].values
    y_income = dataframe[Y].values
    X_train, X_test, y_train, y_test = train_test_split(x_income, y_income, test_size = 0.5, random_state = 35)
    x_income_convert = x_income.reshape(x_income.shape[0], -1) # convert
    regr = LinearRegression().fit(x_income_convert, y_income)
    value_y = regr.predict([[float(self.salary_textedit.toPlainText()) / 10000000]])
    return "Bu???n" if value_y < 2 else ("B??nh th?????ng" if value_y >= 2 and value_y <= 6 else "Vui")

  def show_predict_result(self):
    if (self.is_has_error_input()):
      return
    msg = QtWidgets.QMessageBox()
    message = ""
    msg.setIcon(QtWidgets.QMessageBox.Information)
    message += "H??? t??n: " + self.name_textedit.toPlainText() + "\n" + "Gi???i t??nh: " + self.gender_cb.currentText() + "\n" + "Tu???i: " + self.age_textedit.toPlainText() + "\n" + "Ngh??? nghi???p: " + self.job_textedit.toPlainText() + "\n" + "M???c l????ng: " + self.salary_textedit.toPlainText() + "\n"
    self.output = self.coreProcess()
    message += "K???t qu???: %s" % self.output
    msg.setText(message)
    msg.setWindowTitle("K???t qu??? d??? ??o??n")
    msg.exec_()

  def reset_user_input(self):
    self.name_textedit.setPlainText("")
    self.age_textedit.setPlainText("")
    self.job_textedit.setPlainText("")
    self.salary_textedit.setPlainText("")

  def export_excel(self):

    # create new xlsx instance
    from openpyxl import Workbook
    wb = Workbook()

    export_file_name = "result-export.xlsx"

    data_sheet = wb.create_sheet("Result", 0)
    row_count = data_sheet.max_row

    # dict
    data_inp = {
      "H??? v?? t??n": self.name_textedit.toPlainText(), 
      "Gi???? t??nh": self.gender_cb.currentText(), 
      "Tu???i": self.age_textedit.toPlainText(), 
      "Ngh??? nghi???p": self.job_textedit.toPlainText(), 
      "M???c l????ng": self.salary_textedit.toPlainText(),
      "T??m tr???ng": self.output
    }

    column_names = list(data_inp.keys())

    if row_count == 1:
      for i in range(0, len(column_names)):
        data_sheet.cell(row = row_count, column = i + 1, value = column_names[i])

    # write new data
    for i in range(0, len(column_names)):
      data_sheet.cell(row = row_count + 1, column = i + 1, value = data_inp[column_names[i]])

    wb.save(filename = export_file_name)

    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Information)
    msg.setText("???? ghi th??nh c??ng!")
    msg.setWindowTitle("Th??ng b??o")
    msg.exec_()

if __name__ == "__main__":
  import sys
  app = QtWidgets.QApplication(sys.argv)
  App = QtWidgets.QMainWindow()
  ui = Ui_App()
  ui.setupUi(App)
  App.show()
  sys.exit(app.exec_())
