from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import numpy as np
import pandas as pd
import statistics
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split

class Ui_MainWindow(object):

  def __init__(self):
    self.output = ""

  def setupUi(self, MainWindow):
    MainWindow.setObjectName("MainWindow")
    MainWindow.resize(709, 600)
    MainWindow.setStyleSheet("background-color:rgb(234, 234, 255);\n""")
    self.centralwidget = QtWidgets.QWidget(MainWindow)
    self.centralwidget.setObjectName("centralwidget")
    self.label = QtWidgets.QLabel(self.centralwidget) 
    self.label.setGeometry(QtCore.QRect(90, 30, 571, 51))
    font = QtGui.QFont() 
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label.setFont(font)
    self.label.setStyleSheet(
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 25px;\n"
      "font-weight: bold;\n"
      "color: red;"
    )
    self.label.setObjectName("label")
    self.label_2 = QtWidgets.QLabel(self.centralwidget)
    self.label_2.setGeometry(QtCore.QRect(80, 120, 111, 31))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label_2.setFont(font)
    self.label_2.setStyleSheet(
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 21px;\n"
      "font-weight: bold;"
    )
    self.label_2.setObjectName("label_2")
    self.label_3 = QtWidgets.QLabel(self.centralwidget)
    self.label_3.setGeometry(QtCore.QRect(80, 180, 91, 31))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label_3.setFont(font)
    self.label_3.setStyleSheet(
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 21px;\n"
      "font-weight: bold;"
    )
    self.label_3.setObjectName("label_3")
    self.txthovaten = QtWidgets.QTextEdit(self.centralwidget)
    self.txthovaten.setGeometry(QtCore.QRect(220, 120, 371, 31))
    self.txthovaten.setStyleSheet("background-color: #fff")
    self.txthovaten.setObjectName("txthovaten")
    self.txtgioitinh = QtWidgets.QTextEdit(self.centralwidget)
    self.txtgioitinh.setGeometry(QtCore.QRect(220, 180, 371, 31))
    self.txtgioitinh.setStyleSheet("background-color: #fff")
    self.txtgioitinh.setObjectName("txtgioitinh")
    self.txttuoi = QtWidgets.QTextEdit(self.centralwidget)
    self.txttuoi.setGeometry(QtCore.QRect(220, 240, 371, 31))
    self.txttuoi.setStyleSheet("background-color: #fff")
    self.txttuoi.setObjectName("txttuoi")
    self.label_4 = QtWidgets.QLabel(self.centralwidget)
    self.label_4.setGeometry(QtCore.QRect(80, 240, 55, 31))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label_4.setFont(font)
    self.label_4.setStyleSheet(
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 21px;\n"
      "font-weight: bold;"
    )
    self.label_4.setObjectName("label_4")
    self.label_5 = QtWidgets.QLabel(self.centralwidget)
    self.label_5.setGeometry(QtCore.QRect(80, 290, 131, 41))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label_5.setFont(font)
    self.label_5.setStyleSheet(
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 21px;\n"
      "font-weight: bold;"
    )
    self.label_5.setObjectName("label_5")
    self.txtnghenghiep = QtWidgets.QTextEdit(self.centralwidget)
    self.txtnghenghiep.setGeometry(QtCore.QRect(220, 300, 371, 31))
    self.txtnghenghiep.setStyleSheet("background-color:#fff")
    self.txtnghenghiep.setObjectName("txtnghenghiep")
    self.txtmucluong = QtWidgets.QTextEdit(self.centralwidget)
    self.txtmucluong.setGeometry(QtCore.QRect(220, 360, 371, 31))
    self.txtmucluong.setStyleSheet("background-color:#fff")
    self.txtmucluong.setObjectName("txtmucluong")
    self.label_6 = QtWidgets.QLabel(self.centralwidget)
    self.label_6.setGeometry(QtCore.QRect(80, 350, 111, 41))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.label_6.setFont(font)
    self.label_6.setStyleSheet( 
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size: 21px;\n"
      "font-weight: bold;"
    )
    self.label_6.setObjectName("label_6")
    self.btndudoan = QtWidgets.QPushButton(self.centralwidget)
    self.btndudoan.setGeometry(QtCore.QRect(80, 450, 161, 51))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(-1)
    font.setBold(True)
    font.setWeight(75)
    self.btndudoan.setFont(font)
    self.btndudoan.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
    self.btndudoan.setStyleSheet(
      "background-color: #3c8dbc;\n"
      "border-color: #367fa9;\n"
      "color: white;\n"
      "border-width: 2px;\n"
      "border-radius: 10px;\n"
      "padding: 6px;\n"
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;\n"
      "font-size:21px;\n"
      "font-weight: bold;"
    )
    self.btndudoan.setObjectName("btndudoan")
    self.btnluuvaoexcel = QtWidgets.QPushButton(self.centralwidget)
    self.btnluuvaoexcel.setGeometry(QtCore.QRect(280, 450, 161, 51))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(12)
    font.setBold(True)
    font.setWeight(75)
    self.btnluuvaoexcel.setFont(font)
    self.btnluuvaoexcel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
    self.btnluuvaoexcel.setStyleSheet(
      "background-color: #99CC00;\n"
      "border-color: #367fa9;\n"
      "color: white;\n"
      "border-width: 2px;\n"
      "border-radius: 10px;\n"
      "padding: 6px;\n"
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;"
    )
    self.btnluuvaoexcel.setObjectName("btnluuvaoexcel")
    self.btnnhaplai = QtWidgets.QPushButton(self.centralwidget)
    self.btnnhaplai.setGeometry(QtCore.QRect(470, 450, 161, 51))
    font = QtGui.QFont()
    font.setFamily("Source Sans Pro,Helvetica Neue,Helvetica,Arial,sans-serif")
    font.setPointSize(12)
    font.setBold(True)
    font.setWeight(75)
    self.btnnhaplai.setFont(font)
    self.btnnhaplai.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
    self.btnnhaplai.setStyleSheet(
      "background-color: #800000;\n"
      "border-color: #367fa9;\n"
      "color: white;\n"
      "border-width: 2px;\n"
      "border-radius: 10px;\n"
      "padding: 6px;\n"
      "font-family:\'Source Sans Pro\',\'Helvetica Neue\',Helvetica,Arial,sans-serif;"
    )
    self.btnnhaplai.setObjectName("btnnhaplai")
    MainWindow.setCentralWidget(self.centralwidget)
    self.statusbar = QtWidgets.QStatusBar(MainWindow)
    self.statusbar.setObjectName("statusbar")
    MainWindow.setStatusBar(self.statusbar)

    self.retranslateUi(MainWindow)
    QtCore.QMetaObject.connectSlotsByName(MainWindow)
    self.btndudoan.clicked.connect(self.calcTotal)
    self.btnnhaplai.clicked.connect(self.nhaplai)
    self.btnluuvaoexcel.clicked.connect(self.exportExcel)
    
  def retranslateUi(self, MainWindow):
    _translate = QtCore.QCoreApplication.translate
    MainWindow.setWindowTitle(_translate("MainWindow", "Chương trình"))
    self.label.setText(_translate("MainWindow", "Dự đoán chỉ số hạnh phúc giựa vào mức lương"))
    self.label_2.setText(_translate("MainWindow", "Họ và tên"))
    self.label_3.setText(_translate("MainWindow", "Giới tính"))
    self.label_4.setText(_translate("MainWindow", "Tuổi"))
    self.label_5.setText(_translate("MainWindow", "Nghề nghiệp"))
    self.label_6.setText(_translate("MainWindow", "Mức lương"))
    self.btndudoan.setText(_translate("MainWindow", "Dự đoán"))
    self.btnluuvaoexcel.setText(_translate("MainWindow", "Lưu vào excel"))
    self.btnnhaplai.setText(_translate("MainWindow", "Nhập lại"))

  def coreProcess(self): 
    from sklearn.linear_model import LinearRegression
    dataframe = pd.read_csv("income.data.csv")
    X = "income"
    Y = "happiness"
    x_income = dataframe[X].values
    y_income = dataframe[Y].values
    X_train, X_test, y_train, y_test = train_test_split(x_income, y_income, test_size = 0.5, random_state = 35)
    x_income_convert = x_income.reshape(x_income.shape[0], -1) # convert
    regr = LinearRegression().fit(x_income_convert, y_income)
    value_y = regr.predict([[float(self.txtmucluong.toPlainText()) / 10000000]])
    return "Buồn" if value_y < 2 else ("Bình thường" if value_y >= 2 and value_y <= 6 else "Vui")

  def calcTotal(self):
    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Information)
    message 
      += "Họ tên: " + self.txthovaten.toPlainText() + "\n"
      +  "Giới tính: " + self.txtgioitinh.toPlainText() + "\n"
      +  "Tuổi: " + self.txttuoi.toPlainText() + "\n"
      +  "Nghề nghiệp: " + self.txtnghenghiep.toPlainText() + "\n"
      +  "Mức lương: " + self.txtmucluong.toPlainText() + "\n"

    self.output = self.coreProcess()
    message += "Kết quả: %s" % self.output
    msg.setText(message)
    msg.setWindowTitle("Kết quả dự đoán")
    msg.exec_()

  def nhaplai(self):
    self.txthovaten.setPlainText("")
    self.txtgioitinh.setPlainText("")
    self.txttuoi.setPlainText("")
    self.txtnghenghiep.setPlainText("")
    self.txtmucluong.setPlainText("")
        
  def exportExcel(self):

    # create new xlsx instance
    from openpyxl import Workbook
    wb = Workbook()

    export_file_name = "result-export.xlsx"

    data_sheet = wb.create_sheet(title = "Result")
    row_count = data_sheet.max_row

    # dict
    data_inp = {
      "Họ và tên": self.txthovaten.toPlainText(), 
      "Giơí tính": self.txtgioitinh.toPlainText(), 
      "Tuổi": self.txttuoi.toPlainText(), 
      "Nghề nghiệp": self.txtnghenghiep.toPlainText(), 
      "Mức lương": self.txtmucluong.toPlainText(),
      "Tâm trạng": self.result
    }

    column_names = list(data_inp.keys())

    if row_count == 1:
      for i in range(0, len(column_names)):
        data_sheet.cell(row = row_count, column = i + 1, value = column_names[i])

    # write new data
    for i in range(0, len(column_names)):
      data_sheet.cell(row = row_count + 1, column = i + 1, value = data_inp[column_names[i]])

    wb.save(filename = export_file_name)

    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Information)
    msg.setText("Đã ghi thành công!")
    msg.setWindowTitle("Thông báo")
    msg.exec_()
        
if __name__ == "__main__":
  import sys
  app = QtWidgets.QApplication(sys.argv)
  MainWindow = QtWidgets.QMainWindow()
  ui = Ui_MainWindow()
  ui.setupUi(MainWindow)
  MainWindow.show()
  sys.exit(app.exec_())
